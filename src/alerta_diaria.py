"""Fase 6 — Alerta diaria de producción.

Junta las tres piezas: histórico procesado + modelo entrenado + reporte de
pedidos cargados de las ~11:00, y genera el Excel de llamadas para el
equipo comercial antes del cierre de las 11:59.

Uso (recordar la semántica de fechas: hoy a la mañana se cargan pedidos
con fecha de entrega de MAÑANA; la fecha que se pasa es la de entrega):

    python src/alerta_diaria.py 2026-07-18 data/raw/Pedidos_18-07-2026.xlsx

Salida: outputs/alerta_AAAA-MM-DD.xlsx con la lista priorizada, y el top
15 impreso en pantalla.
"""
import pickle
import sys

import pandas as pd

try:
    import config
    from baseline import construir_pedidos, es_dia_de_alerta, leer_reporte_diario
    from features import FEATURES, construir_panel
except ImportError:
    from src import config
    from src.baseline import (construir_pedidos, es_dia_de_alerta,
                              leer_reporte_diario)
    from src.features import FEATURES, construir_panel

UMBRAL_MODELO = 0.5   # prob. mínima del modelo para entrar en la alerta


def generar_alerta_modelo(fecha, ruta_reporte):
    fecha = pd.Timestamp(fecha)

    # 1) Histórico + features del día objetivo (solo con pasado)
    df = pd.read_parquet(config.VENTAS_PARQUET)
    ped = construir_pedidos(df)
    panel = construir_panel(ped, fecha_prediccion=fecha)
    hoy = panel[panel["Fecha"] == fecha].copy()
    if hoy.empty:
        raise SystemExit("Sin observaciones para esa fecha: revisar histórico.")

    # 2) Modelo entrenado
    with open(config.DATA_PROCESSED / "modelo.pkl", "rb") as f:
        m = pickle.load(f)
    hoy["prob_pedido"] = m["modelo"].predict_proba(hoy[FEATURES].fillna(0))[:, 1]

    # 3) Reporte de las 11:00: quiénes ya cargaron
    lugares_cargados = leer_reporte_diario(ruta_reporte)
    hoy["fantasia"] = hoy["punto_entrega"].str.split(" | ", regex=False).str[0]
    hoy["ya_pidio"] = hoy["fantasia"].isin(lugares_cargados)

    # 4) La alerta: alta probabilidad y todavía no cargado
    alerta = (hoy[(hoy["prob_pedido"] >= UMBRAL_MODELO) & (~hoy["ya_pidio"])]
              .sort_values(["prob_pedido", "cajas_prom"], ascending=False))

    # Contexto de negocio para quien llama
    contacto = (ped.groupby("punto_entrega")
                   .agg(cliente=("cliente", "first"),
                        ultimo_pedido=("Fecha", "max")))
    alerta = alerta.merge(contacto, on="punto_entrega", how="left")
    alerta["tasa_dia_pct"] = (alerta["tasa_dia_semana"] * 100).round(0)
    dia_nombre = ["lunes", "martes", "miércoles", "jueves",
                  "viernes", "sábado", "domingo"][fecha.dayofweek]

    cols = alerta[["cliente", "fantasia", "prob_pedido", "tasa_dia_pct",
                   "ultimo_pedido", "cajas_prom"]].copy()
    cols.columns = ["Cliente (razón social)", "Punto de entrega",
                    "Prob. de pedido", f"% de {dia_nombre}s que pide",
                    "Último pedido", "Cajas promedio"]
    return cols, len(lugares_cargados)


def escribir_excel(tabla: pd.DataFrame, ruta, fecha):
    """Excel con formato simple y legible para el comercial."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(ruta, engine="openpyxl") as xw:
        tabla.to_excel(xw, index=False, sheet_name="Llamar", startrow=1)
        ws = xw.sheets["Llamar"]

        ws["A1"] = (f"Alerta de pedidos — entrega {fecha:%d/%m/%Y} — "
                    f"generada antes del cierre 11:59")
        ws["A1"].font = Font(bold=True, size=12)

        header_fill = PatternFill("solid", fgColor="1F4E79")
        for c in range(1, len(tabla.columns) + 1):
            cell = ws.cell(row=2, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        rojo = PatternFill("solid", fgColor="F8CBAD")     # prob >= 0.8
        amarillo = PatternFill("solid", fgColor="FFE699")  # 0.5 - 0.8
        col_prob = list(tabla.columns).index("Prob. de pedido") + 1
        for r in range(3, len(tabla) + 3):
            p = ws.cell(row=r, column=col_prob)
            p.number_format = "0%"
            fill = rojo if (p.value or 0) >= 0.8 else amarillo
            for c in range(1, len(tabla.columns) + 1):
                ws.cell(row=r, column=c).fill = fill

        col_fecha = list(tabla.columns).index("Último pedido") + 1
        for r in range(3, len(tabla) + 3):
            ws.cell(row=r, column=col_fecha).number_format = "DD/MM/YYYY"

        anchos = [34, 26, 14, 18, 13, 13]
        for i, w in enumerate(anchos[:len(tabla.columns)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        raise SystemExit(1)
    fecha, ruta_reporte = sys.argv[1], sys.argv[2]

    if not es_dia_de_alerta(fecha):
        print(f"{fecha} es domingo o feriado: no hay reparto, no hay alerta.")
        return

    tabla, n_cargados = generar_alerta_modelo(fecha, ruta_reporte)
    config.OUTPUTS.mkdir(exist_ok=True)
    salida = config.OUTPUTS / f"alerta_{fecha}.xlsx"
    escribir_excel(tabla, salida, pd.Timestamp(fecha))

    print(f"Pedidos ya cargados en el reporte: {n_cargados} lugares")
    print(f"Puntos en alerta: {len(tabla)}  ->  {salida}\n")
    print("TOP 15 PARA LLAMAR AHORA:")
    print(tabla.head(15).round(2).to_string(index=False))


if __name__ == "__main__":
    main()
